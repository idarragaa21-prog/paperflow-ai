from __future__ import annotations

import csv
import io
import json
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.logger import logger

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F5496")
_MAX_COL_WIDTH = 60


SHEETS = [
    "STUDIES",
    "ARMS",
    "OUTCOMES",
    "EFFECT_SIZES",
    "RISK_OF_BIAS",
    "TABLES_RAW",
    "IMAGES_OCR_RAW",
    "LOGS",
]


STUDIES_COLS = [
    "study_id",
    "citation_key",
    "title",
    "first_author",
    "year",
    "journal",
    "doi",
    "pmid",
    "pmcid",
    "country",
    "centers",
    "study_design",
    "recruitment_dates",
    "follow_up_duration",
    "population_description",
    "inclusion_criteria",
    "exclusion_criteria",
    "baseline_balance_notes",
    "funding",
    "conflicts_of_interest",
    "registration",
    "notes",
    "total_sample_size",
    "total_n_intervention",
    "total_n_control",
    "mean_age_total",
    "percent_female_total",
    "setting",
    "geographic_region",
    "income_level",
    "language_of_publication",
    "peer_reviewed",
    "extraction_confidence",
    "provenance",
    "raw_study_json",
]

ARMS_COLS = [
    "arm_id",
    "study_id",
    "arm_name",
    "intervention_description",
    "dose_intensity",
    "cointerventions",
    "n_randomized",
    "n_analyzed",
    "losses_to_followup_n",
    "losses_to_followup_reasons",
    "mean_age",
    "sd_age",
    "median_age",
    "iqr_age",
    "age_range",
    "percent_female",
    "percent_male",
    "n_female",
    "n_male",
    "mean_bmi",
    "sd_bmi",
    "comorbidities",
    "disease_severity",
    "disease_duration",
    "baseline_outcome_value",
    "previous_treatments",
    "smoking_status",
    "ethnicity_distribution",
    "followup_duration_months",
    "completion_rate",
    "withdrawal_reasons",
]

OUTCOMES_COLS = [
    "outcome_id",
    "study_id",
    "outcome_name",
    "outcome_type",
    "definition",
    "timepoint",
    "measurement_instrument",
    "outcome_category",
    "direction_of_benefit",
    "how_measured",
    "measurement_timepoints",
    "missing_data_handling",
]

EFFECTS_COLS = [
    "effect_id",
    "study_id",
    "outcome_name",
    "timepoint",
    "arm_intervention",
    "arm_control",
    "effect_measure",
    "a_events",
    "b_non_events",
    "c_events",
    "d_non_events",
    "continuity_correction_used",
    "or_value",
    "log_or",
    "se_log_or",
    "ci_lower_95",
    "ci_upper_95",
    "adjusted_or",
    "adjusted_rr",
    "adjusted_hr",
    "adjustment_variables",
    "is_adjusted",
    "raw_extracted_value",
    "page_number",
    "table_id",
    "figure_id",
    "source_type",
    "extraction_confidence",
    "comments",
    "manually_edited",
    "edited_at",
    "p_value",
    "n_intervention_for_outcome",
    "n_control_for_outcome",
    "events_intervention",
    "events_control",
    "mean_intervention",
    "sd_intervention",
    "mean_control",
    "sd_control",
    "median_intervention",
    "iqr_intervention",
    "median_control",
    "iqr_control",
    "mean_difference",
    "sd_difference",
    "nnt",
    "absolute_risk_reduction",
    "relative_risk_reduction",
    "heterogeneity_subgroup",
    "sensitivity_analysis",
    "analysis_method",
    "model_type",
    "intention_to_treat",
]

ROB_COLS = [
    "rob_id",
    "study_id",
    "tool",
    "domain_name",
    "judgement",
    "support_for_judgement",
    "manually_edited",
    "edited_at",
]

TABLES_RAW_COLS = ["table_id", "study_id", "page", "extracted_markdown"]
IMAGES_OCR_RAW_COLS = ["image_id", "study_id", "page", "ocr_text"]
LOGS_COLS = ["timestamp", "job_id", "study_id", "level", "message"]


@dataclass
class ExcelExportInput:
    studies: list[dict[str, Any]]
    arms: list[dict[str, Any]]
    outcomes: list[dict[str, Any]]
    effects: list[dict[str, Any]]
    rob: list[dict[str, Any]]
    tables_raw: list[dict[str, Any]]
    images_ocr_raw: list[dict[str, Any]]
    logs: list[dict[str, Any]]


def _add_sheet(wb: Workbook, name: str, cols: list[str]) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(cols)
    # Bold + colored header row
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    # Freeze header row and enable autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _cell_value(v: Any) -> Any:
    """Convert non-primitive values to JSON strings so openpyxl can write them."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return json.dumps(v, ensure_ascii=False)


def _append_rows(ws, cols: list[str], rows: list[dict[str, Any]]) -> None:
    for r in rows:
        ws.append([_cell_value(r.get(c)) for c in cols])


def _autofit_columns(ws) -> None:
    """Set column widths based on content, capped at _MAX_COL_WIDTH."""
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_len = len(str(cell.value))
                col_widths[cell.column] = min(
                    _MAX_COL_WIDTH,
                    max(col_widths.get(cell.column, 12), val_len + 2),
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_meta_xlsx(data: ExcelExportInput) -> bytes:
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    # Create sheets
    _add_sheet(wb, "STUDIES", STUDIES_COLS)
    _add_sheet(wb, "ARMS", ARMS_COLS)
    _add_sheet(wb, "OUTCOMES", OUTCOMES_COLS)
    _add_sheet(wb, "EFFECT_SIZES", EFFECTS_COLS)
    _add_sheet(wb, "RISK_OF_BIAS", ROB_COLS)
    _add_sheet(wb, "TABLES_RAW", TABLES_RAW_COLS)
    _add_sheet(wb, "IMAGES_OCR_RAW", IMAGES_OCR_RAW_COLS)
    _add_sheet(wb, "LOGS", LOGS_COLS)

    # Fill
    _append_rows(wb["STUDIES"], STUDIES_COLS, data.studies)
    _append_rows(wb["ARMS"], ARMS_COLS, data.arms)
    _append_rows(wb["OUTCOMES"], OUTCOMES_COLS, data.outcomes)
    _append_rows(wb["EFFECT_SIZES"], EFFECTS_COLS, data.effects)
    _append_rows(wb["RISK_OF_BIAS"], ROB_COLS, data.rob)
    _append_rows(wb["TABLES_RAW"], TABLES_RAW_COLS, data.tables_raw)
    _append_rows(wb["IMAGES_OCR_RAW"], IMAGES_OCR_RAW_COLS, data.images_ocr_raw)
    _append_rows(wb["LOGS"], LOGS_COLS, data.logs)

    # Auto-fit column widths on all sheets
    for sheet_name in wb.sheetnames:
        _autofit_columns(wb[sheet_name])

    from io import BytesIO

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── CSV bundle export ────────────────────────────────────────────────────────

_SHEET_MAP = [
    ("studies", STUDIES_COLS, "studies"),
    ("arms", ARMS_COLS, "arms"),
    ("outcomes", OUTCOMES_COLS, "outcomes"),
    ("effects", EFFECTS_COLS, "effect_sizes"),
    ("rob", ROB_COLS, "risk_of_bias"),
    ("tables_raw", TABLES_RAW_COLS, "tables_raw"),
    ("images_ocr_raw", IMAGES_OCR_RAW_COLS, "images_ocr_raw"),
    ("logs", LOGS_COLS, "logs"),
]


def _rows_to_csv(cols: list[str], rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(cols)
    for r in rows:
        writer.writerow([r.get(c) for c in cols])
    return buf.getvalue()


def build_meta_csv_bundle(data: ExcelExportInput) -> bytes:
    """Return a ZIP archive containing one CSV per sheet plus a manifest.json."""
    data_map: dict[str, list[dict[str, Any]]] = {
        "studies": data.studies,
        "arms": data.arms,
        "outcomes": data.outcomes,
        "effects": data.effects,
        "rob": data.rob,
        "tables_raw": data.tables_raw,
        "images_ocr_raw": data.images_ocr_raw,
        "logs": data.logs,
    }

    manifest: dict[str, Any] = {
        "format": "csv_bundle",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "tables": {},
    }

    bio = io.BytesIO()
    with zipfile.ZipFile(bio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for attr, cols, filename in _SHEET_MAP:
            rows = data_map[attr]
            csv_text = _rows_to_csv(cols, rows)
            zf.writestr(f"{filename}.csv", csv_text)
            manifest["tables"][filename] = {
                "filename": f"{filename}.csv",
                "columns": cols,
                "rows": len(rows),
            }

        zf.writestr("manifest.json", json.dumps(manifest, indent=2))

    return bio.getvalue()
