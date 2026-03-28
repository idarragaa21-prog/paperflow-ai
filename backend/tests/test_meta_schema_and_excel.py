from __future__ import annotations

import json
import zipfile

import pytest
from openpyxl import load_workbook

from app.services.meta_extractor.excel_export import (
    ARMS_COLS,
    EFFECTS_COLS,
    IMAGES_OCR_RAW_COLS,
    LOGS_COLS,
    OUTCOMES_COLS,
    ROB_COLS,
    STUDIES_COLS,
    TABLES_RAW_COLS,
    ExcelExportInput,
    build_meta_csv_bundle,
    build_meta_xlsx,
)
from app.services.meta_extractor.schema import EffectSizeSchema, ExtractedStudySchema
from app.services.meta_extractor.quality_gate import quality_gate


def test_meta_schema_validates():
    data = {
        "title": "T",
        "first_author": "A",
        "year": 2024,
        "journal": "J",
        "arms": [{"arm_name": "Intervention"}, {"arm_name": "Control"}],
        "outcomes": [{"outcome_name": "Mortality"}],
        "effect_sizes": [],
        "risk_of_bias": [],
    }
    ExtractedStudySchema.model_validate(data)


def test_meta_schema_allows_null_2x2_when_adjusted():
    es = EffectSizeSchema(
        outcome_name="Mortality",
        arm_intervention="Intervention",
        arm_control="Control",
        effect_measure="OR",
        is_adjusted=True,
        adjusted_or=1.2,
        a_events=None,
        b_non_events=None,
        c_events=None,
        d_non_events=None,
    )
    assert es.is_adjusted is True


def test_meta_schema_rejects_negative_events():
    with pytest.raises(Exception):
        EffectSizeSchema(
            outcome_name="Mortality",
            arm_intervention="Intervention",
            arm_control="Control",
            a_events=-1,
        )


def test_quality_gate_recalculates_or_and_detects_mismatch():
    study = ExtractedStudySchema.model_validate(
        {
            "title": "T",
            "first_author": "A",
            "year": 2024,
            "journal": "J",
            "arms": [{"arm_name": "Intervention", "n_randomized": 10, "n_analyzed": 10}, {"arm_name": "Control", "n_randomized": 10, "n_analyzed": 10}],
            "outcomes": [{"outcome_name": "Mortality"}],
            "effect_sizes": [
                {
                    "outcome_name": "Mortality",
                    "arm_intervention": "Intervention",
                    "arm_control": "Control",
                    "a_events": 2,
                    "b_non_events": 8,
                    "c_events": 1,
                    "d_non_events": 9,
                    "or_value": 99.0,
                }
            ],
            "risk_of_bias": [],
        }
    )

    fixed, warnings = quality_gate(study)
    assert fixed.effect_sizes[0].or_value is not None
    assert any("OR mismatch" in w for w in warnings)
    assert fixed.effect_sizes[0].log_or is not None
    assert fixed.effect_sizes[0].se_log_or is not None


def test_excel_export_has_all_sheets_and_column_order():
    xlsx = build_meta_xlsx(
        ExcelExportInput(
            studies=[],
            arms=[],
            outcomes=[],
            effects=[],
            rob=[],
            tables_raw=[],
            images_ocr_raw=[],
            logs=[],
        )
    )
    wb = load_workbook(filename=bytes_to_filelike(xlsx))
    assert wb.sheetnames == [
        "STUDIES",
        "ARMS",
        "OUTCOMES",
        "EFFECT_SIZES",
        "RISK_OF_BIAS",
        "TABLES_RAW",
        "IMAGES_OCR_RAW",
        "LOGS",
    ]

    assert [c.value for c in next(wb["STUDIES"].iter_rows(max_row=1))] == STUDIES_COLS
    assert [c.value for c in next(wb["ARMS"].iter_rows(max_row=1))] == ARMS_COLS
    assert [c.value for c in next(wb["OUTCOMES"].iter_rows(max_row=1))] == OUTCOMES_COLS
    assert [c.value for c in next(wb["EFFECT_SIZES"].iter_rows(max_row=1))] == EFFECTS_COLS
    assert [c.value for c in next(wb["RISK_OF_BIAS"].iter_rows(max_row=1))] == ROB_COLS
    assert [c.value for c in next(wb["TABLES_RAW"].iter_rows(max_row=1))] == TABLES_RAW_COLS
    assert [c.value for c in next(wb["IMAGES_OCR_RAW"].iter_rows(max_row=1))] == IMAGES_OCR_RAW_COLS
    assert [c.value for c in next(wb["LOGS"].iter_rows(max_row=1))] == LOGS_COLS


def test_excel_export_includes_manual_edit_flags():
    xlsx = build_meta_xlsx(
        ExcelExportInput(
            studies=[],
            arms=[],
            outcomes=[],
            effects=[
                {
                    "effect_id": "E1",
                    "study_id": "S1",
                    "outcome_name": "Mortality",
                    "arm_intervention": "I",
                    "arm_control": "C",
                    "effect_measure": "OR",
                    "a_events": 2,
                    "b_non_events": 8,
                    "c_events": 1,
                    "d_non_events": 9,
                    "manually_edited": True,
                    "edited_at": "2026-02-09T00:00:00Z",
                }
            ],
            rob=[
                {
                    "rob_id": "R1",
                    "study_id": "S1",
                    "tool": "ROB2",
                    "domain_name": "randomization",
                    "judgement": "low",
                    "support_for_judgement": "ok",
                    "manually_edited": True,
                    "edited_at": "2026-02-09T00:00:00Z",
                }
            ],
            tables_raw=[],
            images_ocr_raw=[],
            logs=[],
        )
    )
    wb = load_workbook(filename=bytes_to_filelike(xlsx))

    ws = wb["EFFECT_SIZES"]
    # header + 1 row
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == tuple(EFFECTS_COLS)
    # map row by col
    row = {col: rows[1][i] for i, col in enumerate(EFFECTS_COLS)}
    assert row["manually_edited"] is True
    assert row["edited_at"] == "2026-02-09T00:00:00Z"

    ws2 = wb["RISK_OF_BIAS"]
    rows2 = list(ws2.iter_rows(values_only=True))
    assert rows2[0] == tuple(ROB_COLS)
    row2 = {col: rows2[1][i] for i, col in enumerate(ROB_COLS)}
    assert row2["manually_edited"] is True
    assert row2["edited_at"] == "2026-02-09T00:00:00Z"


def test_csv_bundle_export_includes_all_metadata_tables_and_manifest():
    bundle = build_meta_csv_bundle(
        ExcelExportInput(
            studies=[{"study_id": "S1", "title": "Study 1"}],
            arms=[{"arm_id": "A1", "study_id": "S1", "arm_name": "Intervention"}],
            outcomes=[{"outcome_id": "O1", "study_id": "S1", "outcome_name": "Pain"}],
            effects=[{"effect_id": "E1", "study_id": "S1", "outcome_name": "Pain"}],
            rob=[{"rob_id": "R1", "study_id": "S1", "tool": "ROB2"}],
            tables_raw=[],
            images_ocr_raw=[],
            logs=[],
        )
    )

    with zipfile.ZipFile(bytes_to_filelike(bundle)) as zf:
        names = set(zf.namelist())
        assert {
            "manifest.json",
            "studies.csv",
            "arms.csv",
            "outcomes.csv",
            "effect_sizes.csv",
            "risk_of_bias.csv",
            "tables_raw.csv",
            "images_ocr_raw.csv",
            "logs.csv",
        }.issubset(names)

        manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
        assert manifest["format"] == "csv_bundle"
        assert manifest["tables"]["studies"]["rows"] == 1

        studies_csv = zf.read("studies.csv").decode("utf-8")
        assert studies_csv.splitlines()[0].split(",") == STUDIES_COLS
        assert "Study 1" in studies_csv


def bytes_to_filelike(data: bytes):
    import io

    return io.BytesIO(data)
