"""Comprehensive extraction (fan-out) + Excel data-sheet building."""
from __future__ import annotations

import io

import metaforge.extract as ex
from metaforge.excel_export import extractions_to_xlsx


def _parts():
    return {
        "narrative": "Ensayo aleatorizado en España. La mortalidad fue menor con el fármaco.",
        "tables": [
            "Tabla 1 Basales | Edad | 60 | 61",
            "Tabla 2 Resultados | Muertes int 10/100 | Muertes ctrl 20/100",
        ],
        "captions": ["Figura 1. Curva de supervivencia."],
        "has_fulltext": True,
    }


def _fake_ai(prompt, timeout=150, model=None):
    if "DOS cosas a la vez" in prompt:  # merged study + narrative call
        return {
            "study": {"authors": "Smith", "year": "2020", "country": "España", "design": "ECA"},
            "population": {"description": "adultos", "total_n": 200, "age": "60"},
            "arms": [{"name": "Fármaco", "n": 100}, {"name": "Placebo", "n": 100}],
            "followup": "12 meses",
            # a partial echo of the table outcome, reported in prose (should dedup away)
            "outcomes": [{"name": "Mortalidad", "type": "binary", "timepoint": "12m",
                          "intervention": {"events": 10}, "control": {}, "effect": {}, "source": "texto"}],
        }
    if "TABLA 1" in prompt:
        return {"outcomes": []}  # baseline table → nothing
    if "TABLA 2" in prompt:
        return {"outcomes": [{
            "name": "Mortalidad", "type": "binary", "timepoint": "12m",
            "intervention": {"events": 10, "n": 100}, "control": {"events": 20, "n": 100},
            "effect": {"measure": "OR", "value": 0.44, "ci_lower": 0.2, "ci_upper": 0.98},
            "source": "tabla 2"}]}
    return {"outcomes": []}


def _patch(monkeypatch):
    monkeypatch.setattr(ex, "fetch_fulltext_full", lambda rec, **k: _parts())
    monkeypatch.setattr(ex, "ai_available", lambda: True)
    monkeypatch.setattr(ex, "run_claude_json", _fake_ai)


def test_extract_full_fans_out_and_merges(monkeypatch):
    _patch(monkeypatch)
    out = ex.extract_full({"authors": "Smith", "year": 2020, "pmcid": "PMC1", "doi": "10.x"},
                          mode="ai", max_workers=4)
    assert out["found"] is True
    assert out["note"] == ""
    d = out["data"]
    assert d["study"]["country"] == "España"
    assert d["population"]["total_n"] == 200
    assert len(d["arms"]) == 2
    # The partial narrative echo must be collapsed → exactly one outcome.
    assert len(d["outcomes"]) == 1
    o = d["outcomes"][0]
    assert o["intervention"]["events"] == 10 and o["control"]["events"] == 20
    assert o["effect"]["measure"] == "OR"


def test_extract_full_keeps_partial_on_error(monkeypatch):
    monkeypatch.setattr(ex, "fetch_fulltext_full", lambda rec, **k: _parts())
    monkeypatch.setattr(ex, "ai_available", lambda: True)

    def flaky(prompt, timeout=150, model=None):
        if "TABLA 1" in prompt:
            raise RuntimeError("timeout")  # one section fails
        return _fake_ai(prompt, timeout=timeout, model=model)

    monkeypatch.setattr(ex, "run_claude_json", flaky)
    out = ex.extract_full({"authors": "Smith", "year": 2020, "pmcid": "PMC1"}, mode="ai")
    # Other sections still produce data; the failure is reported but not fatal.
    assert out["found"] is True
    assert len(out["data"]["outcomes"]) == 1
    assert "parcial" in out["note"]


def test_extract_full_without_text():
    out = ex.extract_full({"authors": "X", "year": 2021}, mode="local")
    assert out["found"] is False
    assert out["data"] is None


def test_dedup_keeps_distinct_full_analyses():
    def mk(measure, value, events=None, n=None):
        return {"name": "Mort", "timepoint": "12m",
                "intervention": {"events": events, "n": n, "mean": None, "sd": None, "person_time": None},
                "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "effect": {"measure": measure, "value": value, "ci_lower": value and value - 0.1,
                           "ci_upper": value and value + 0.1, "p": 0.01}}
    full_or = mk("OR", 0.44, events=10, n=100)
    full_or["control"] = {"events": 20, "n": 100, "mean": None, "sd": None, "person_time": None}
    adj_hr = mk("HR", 0.35)
    partial = {"name": "Mort", "timepoint": "12m",
               "intervention": {"events": 10, "n": None, "mean": None, "sd": None, "person_time": None},
               "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
               "effect": {"measure": "", "value": None, "ci_lower": None, "ci_upper": None, "p": None}}
    kept = ex._dedup_outcomes([full_or, partial, adj_hr])
    measures = sorted(o["effect"]["measure"] for o in kept)
    assert measures == ["HR", "OR"]  # both full analyses kept, partial dropped


def test_excel_three_sheets_and_meta_mapping(monkeypatch):
    _patch(monkeypatch)
    out = ex.extract_full({"authors": "Smith", "year": 2020, "pmcid": "PMC1", "doi": "10.x"}, mode="ai")
    xlsx = extractions_to_xlsx([out])
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["Estudios", "Desenlaces", "Para meta-análisis"]
    meta = wb["Para meta-análisis"]
    rows = list(meta.iter_rows(values_only=True))
    assert rows[0][3] == "effect_measure"
    data_rows = [r for r in rows[1:] if r[0]]
    assert len(data_rows) == 1
    r = data_rows[0]
    # OR mapping: a=10,b=90,c=20,d=80
    assert (r[4], r[5], r[6], r[7]) == (10, 90, 20, 80)


def test_meta_measure_normalisation():
    from metaforge.excel_export import _meta_row, _norm_measure
    assert _norm_measure("aHR", "time-to-event") == "HR"
    assert _norm_measure("adjusted hazard ratio", "") == "HR"
    assert _norm_measure("Rate ratio", "") == "IRR"
    assert _norm_measure("aOR", "") == "OR"
    assert _norm_measure("", "time-to-event") == "HR"
    assert _norm_measure("nonsense", "") == "GEN"
    # an outcome carrying only an adjusted HR + CI maps to a valid HR row
    o = {"name": "CRC", "type": "time-to-event", "timepoint": "",
         "intervention": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "effect": {"measure": "aHR", "value": 0.35, "ci_lower": 0.24, "ci_upper": 0.52, "p": None}}
    row = _meta_row("Huang 2026", o)
    assert row["effect_measure"] == "HR"
    assert row["effect_value"] == 0.35 and row["ci_lower_95"] == 0.24


def test_meta_row_surfaces_effect_without_full_ci():
    from metaforge.excel_export import _meta_row
    o = {"name": "SLR", "type": "time-to-event", "timepoint": "",
         "intervention": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "effect": {"measure": "HR", "value": 0.6, "ci_lower": None, "ci_upper": None, "p": 0.04}}
    row = _meta_row("X 2020", o)
    assert row["effect_measure"] == "HR"
    assert row["effect_value"] == 0.6
    assert row["ci_lower_95"] is None  # left blank for manual completion


def test_meta_rows_for_filters_and_maps():
    from metaforge.excel_export import meta_rows_for
    data = {"label": "Z 2025", "study": {"year": "2025"}, "outcomes": [
        {"name": "Mort", "type": "binary", "timepoint": "",
         "intervention": {"events": 10, "n": 100}, "control": {"events": 20, "n": 100},
         "effect": {"measure": "OR", "value": None, "ci_lower": None, "ci_upper": None, "p": None},
         "from_figure": False, "approx": False},
        # estimated-from-figure → excluded
        {"name": "Curva", "type": "time-to-event", "timepoint": "5a",
         "intervention": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
         "effect": {"measure": "HR", "value": 0.4, "ci_lower": None, "ci_upper": None, "p": None},
         "from_figure": True, "approx": True},
        # no usable data → excluded
        {"name": "Vacío", "type": "other", "timepoint": "",
         "intervention": {}, "control": {}, "effect": {"measure": "", "value": None}}]}
    rows = meta_rows_for(data)
    assert len(rows) == 1
    r = rows[0]
    assert r["study_label"] == "Z 2025" and r["year"] == "2025"
    assert r["effect_measure"] == "OR" and r["a_events"] == 10 and r["b_non_events"] == 90


def test_extract_full_includes_meta_rows(monkeypatch):
    import metaforge.extract as ex
    parts = {"narrative": "t", "tables": ["Tabla 2 | x"], "captions": [], "xml": "", "has_fulltext": True}
    monkeypatch.setattr(ex, "fetch_fulltext_full", lambda rec, **k: parts)
    monkeypatch.setattr(ex, "ai_available", lambda: True)

    def fake(prompt, timeout=150, model=None, cwd=None):
        if "DOS cosas a la vez" in prompt:
            return {"study": {"year": "2025"}, "population": {}, "arms": [], "followup": "", "outcomes": []}
        return {"outcomes": [{"name": "Mort", "type": "binary", "timepoint": "",
                              "intervention": {"events": 8, "n": 100}, "control": {"events": 16, "n": 100},
                              "effect": {"measure": "OR"}}]}

    monkeypatch.setattr(ex, "run_claude_json", fake)
    out = ex.extract_full({"pmcid": "PMC1", "authors": "A", "year": 2025}, mode="ai", read_figures=False)
    assert out["meta_rows"] and out["meta_rows"][0]["effect_measure"] == "OR"
    assert out["meta_rows"][0]["a_events"] == 8


def test_excel_handles_empty_extraction():
    xlsx = extractions_to_xlsx([{"data": None, "study_label": "X 2020"}])
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb["Desenlaces"].max_row == 1  # header only
