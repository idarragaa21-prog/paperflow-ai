"""Figure-image reading: JATS parsing, flagging, dedup priority, Excel routing."""
from __future__ import annotations

import io

import metaforge.extract as ex
import metaforge.figvision as fv
from metaforge.excel_export import extractions_to_xlsx

_JATS = """<article xmlns:xlink="http://www.w3.org/1999/xlink">
  <fig id="Fig1"><label>Fig. 1</label>
    <caption><p>Diagrama de flujo del estudio</p></caption>
    <alternatives>
      <graphic content-type="image" xlink:href="art_Fig1_HTML.jpg"/>
      <graphic content-type="thumb" xlink:href="art_Fig1_HTML.gif"/>
    </alternatives>
  </fig>
  <fig id="Fig2"><label>Fig. 2</label>
    <caption><p>Curvas de incidencia acumulada</p></caption>
    <graphic xlink:href="art_Fig2_HTML.jpg"/>
  </fig>
</article>"""


def test_figs_from_xml_picks_full_size_images():
    figs = fv._figs_from_xml(_JATS)
    assert [f["file"] for f in figs] == ["art_Fig1_HTML.jpg", "art_Fig2_HTML.jpg"]
    assert figs[0]["label"] == "Fig. 1"
    assert "flujo" in figs[0]["caption"]


def test_extract_from_figures_flags_and_tags(monkeypatch):
    def fake(prompt, timeout=150, model=None, cwd=None):
        # Fig1 (flowchart) → nothing; Fig2 → a quantitative read
        if "Fig. 2" in prompt:
            return {"outcomes": [{"name": "Incidencia CRC", "type": "time-to-event",
                                  "timepoint": "10 años", "effect": {"measure": "", "value": None},
                                  "intervention": {"events": None}, "control": {}, "notes": ""}]}
        return {"outcomes": []}

    monkeypatch.setattr(fv, "run_claude_json", fake)
    figs = [{"label": "Fig. 1", "caption": "flujo", "path": "/tmp/f1.jpg"},
            {"label": "Fig. 2", "caption": "curvas", "path": "/tmp/f2.jpg"}]
    outs = fv.extract_from_figures(figs, max_workers=2)
    assert len(outs) == 1
    o = outs[0]
    assert o["from_figure"] is True
    assert o["source"] == "Fig. 2"
    assert "aproximado" in o["notes"]


def test_dedup_prefers_exact_over_figure():
    def mk(value, from_figure, events=None, n=None):
        return {"name": "CRC", "timepoint": "10a",
                "intervention": {"events": events, "n": n, "mean": None, "sd": None, "person_time": None},
                "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "effect": {"measure": "HR", "value": value, "ci_lower": None, "ci_upper": None, "p": None},
                "from_figure": from_figure}
    exact = mk(0.35, False, events=42, n=610)
    approx = mk(0.30, True)
    kept = ex._dedup_outcomes([approx, exact])
    assert len(kept) == 1
    assert kept[0]["from_figure"] is False  # the exact one wins
    # a figure read with NO exact counterpart survives (fills a gap)
    only_fig = mk(0.30, True)
    only_fig["timepoint"] = "5a"
    kept2 = ex._dedup_outcomes([exact, only_fig])
    assert {o["timepoint"] for o in kept2} == {"10a", "5a"}


def test_excel_routes_figure_reads_to_desenlaces_only():
    ext = {"study_label": "X 2026", "data": {"label": "X 2026", "study": {"year": "2026"},
           "population": {}, "arms": [], "followup": "", "data_in_figures_only": "", "notes": "",
           "outcomes": [
               {"name": "CRC exacto", "type": "time-to-event", "timepoint": "10a",
                "intervention": {"events": 42, "n": 610, "mean": None, "sd": None, "person_time": None},
                "control": {"events": 88, "n": 598, "mean": None, "sd": None, "person_time": None},
                "effect": {"measure": "HR", "value": 0.46, "ci_lower": 0.32, "ci_upper": 0.66, "p": None},
                "source": "tabla 3", "notes": "", "from_figure": False},
               {"name": "CRC de curva", "type": "time-to-event", "timepoint": "5a",
                "intervention": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "effect": {"measure": "HR", "value": 0.4, "ci_lower": None, "ci_upper": None, "p": None},
                "source": "Fig. 2", "notes": "aproximado", "from_figure": True, "approx": True}]}}
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(extractions_to_xlsx([ext])))
    des = list(wb["Desenlaces"].iter_rows(values_only=True))
    origins = [r[19] for r in des[1:]]  # "Origen" column
    assert "figura: estimado de gráfica (aprox.)" in origins and "texto/tabla" in origins
    # meta sheet: only the exact outcome (the approximate curve read is excluded)
    meta = [r for r in wb["Para meta-análisis"].iter_rows(values_only=True)][1:]
    meta = [r for r in meta if r[0]]
    assert len(meta) == 1
    assert meta[0][2] == "CRC exacto"


def test_excel_includes_printed_figure_values_in_meta():
    # a value PRINTED on a figure (e.g. a subgroup forest-plot aHR) is exact and
    # belongs in the synthesis sheet, flagged by provenance.
    ext = {"study_label": "Y 2026", "data": {"label": "Y 2026", "study": {"year": "2026"},
           "population": {}, "arms": [], "followup": "", "data_in_figures_only": "", "notes": "",
           "outcomes": [
               {"name": "CRC subgrupo varones", "type": "time-to-event", "timepoint": "",
                "intervention": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "control": {"events": None, "n": None, "mean": None, "sd": None, "person_time": None},
                "effect": {"measure": "HR", "value": 0.25, "ci_lower": 0.15, "ci_upper": 0.42, "p": None},
                "source": "Fig. 4", "notes": "valor impreso", "from_figure": True, "approx": False}]}}
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(extractions_to_xlsx([ext])))
    des = [r for r in wb["Desenlaces"].iter_rows(values_only=True)][1:]
    assert des[0][19] == "figura: valor impreso"
    meta = [r for r in wb["Para meta-análisis"].iter_rows(values_only=True)][1:]
    meta = [r for r in meta if r[0]]
    assert len(meta) == 1 and meta[0][3] == "HR" and meta[0][18] == 0.25


def test_extract_full_adds_figure_outcomes(monkeypatch):
    parts = {"narrative": "texto", "tables": ["Tabla 2 | a | b"], "captions": [],
             "xml": "<article/>", "has_fulltext": True}
    monkeypatch.setattr(ex, "fetch_fulltext_full", lambda rec, **k: parts)
    monkeypatch.setattr(ex, "ai_available", lambda: True)

    def fake(prompt, timeout=150, model=None):
        if "DOS cosas a la vez" in prompt:
            return {"study": {"year": "2026"}, "population": {}, "arms": [], "followup": "",
                    "outcomes": []}
        return {"outcomes": [{"name": "Mortalidad", "type": "binary", "timepoint": "",
                              "intervention": {"events": 5, "n": 100}, "control": {"events": 9, "n": 100},
                              "effect": {"measure": "OR"}}]}

    monkeypatch.setattr(ex, "run_claude_json", fake)
    # stub the figure pipeline
    monkeypatch.setattr(fv, "fetch_figure_images",
                        lambda rec, xml, dest: [{"label": "Fig. 2", "caption": "c", "path": "/x.jpg"}])
    monkeypatch.setattr(fv, "extract_from_figures",
                        lambda figs, **k: [{"name": "Incidencia (curva)", "type": "time-to-event",
                                            "timepoint": "10a", "intervention": {}, "control": {},
                                            "effect": {"measure": "HR", "value": 0.3, "ci_lower": 0.2,
                                                       "ci_upper": 0.5},
                                            "from_figure": True, "source": "Fig. 2",
                                            "notes": "aproximado"}])
    out = ex.extract_full({"pmcid": "PMC1", "authors": "A", "year": 2026}, mode="ai")
    names = {o["name"] for o in out["data"]["outcomes"]}
    assert "Mortalidad" in names and "Incidencia (curva)" in names
    assert out["n_figures"] == 1
    fig_out = next(o for o in out["data"]["outcomes"] if o["name"] == "Incidencia (curva)")
    assert fig_out["from_figure"] is True
