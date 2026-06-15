"""Build a comprehensive Excel (.xlsx) data-extraction workbook.

Takes the full extractions (one per study) and writes three sheets:
  * Estudios       — one row per study (characteristics, population, arms).
  * Desenlaces     — one row per study × outcome (all reported numbers).
  * Para meta-análisis — outcomes mapped to MetaForge's columns, ready to paste
                          back into the Data step and synthesise.
Nothing is invented: blank cells mean the value was not reported (or only in a
figure image, which is flagged in the Estudios sheet).
"""
from __future__ import annotations

import io


def _autosize(ws, widths: dict[int, int]) -> None:
    from openpyxl.utils import get_column_letter
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="EEF0FB")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if headers:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _arm_lookup(outcome: dict, arm: str) -> dict:
    return outcome.get(arm) or {}


def _meta_row(study_label: str, o: dict) -> dict:
    """Map one extracted outcome to MetaForge's analysis columns where possible."""
    iv, ct = _arm_lookup(o, "intervention"), _arm_lookup(o, "control")
    eff = o.get("effect") or {}
    typ = (o.get("type") or "").lower()
    row = {"study_label": study_label, "year": "", "outcome": o.get("name", ""), "effect_measure": ""}
    # binary: events + n in both arms -> 2x2
    if iv.get("events") is not None and iv.get("n") and ct.get("events") is not None and ct.get("n"):
        row.update({"effect_measure": "OR", "a_events": iv["events"], "b_non_events": iv["n"] - iv["events"],
                    "c_events": ct["events"], "d_non_events": ct["n"] - ct["events"]})
    # continuous: means/SDs/n in both arms
    elif None not in (iv.get("mean"), iv.get("sd"), iv.get("n"), ct.get("mean"), ct.get("sd"), ct.get("n")):
        row.update({"effect_measure": "SMD", "n_intervention": iv["n"], "mean_intervention": iv["mean"],
                    "sd_intervention": iv["sd"], "n_control": ct["n"], "mean_control": ct["mean"], "sd_control": ct["sd"]})
    # person-time -> IRR
    elif iv.get("events") is not None and iv.get("person_time") and ct.get("events") is not None and ct.get("person_time"):
        row.update({"effect_measure": "IRR", "events_intervention": iv["events"], "time_intervention": iv["person_time"],
                    "events_control": ct["events"], "time_control": ct["person_time"]})
    # precomputed effect + CI
    elif eff.get("value") is not None and eff.get("ci_lower") is not None and eff.get("ci_upper") is not None:
        m = (eff.get("measure") or ("HR" if "time" in typ else "OR")).upper()
        row.update({"effect_measure": m, "effect_value": eff["value"],
                    "ci_lower_95": eff["ci_lower"], "ci_upper_95": eff["ci_upper"]})
    return row


_META_COLS = ["study_label", "year", "outcome", "effect_measure",
              "a_events", "b_non_events", "c_events", "d_non_events",
              "n_intervention", "mean_intervention", "sd_intervention",
              "n_control", "mean_control", "sd_control",
              "events_intervention", "time_intervention", "events_control", "time_control",
              "effect_value", "ci_lower_95", "ci_upper_95"]


def extractions_to_xlsx(extractions: list[dict]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    # ---- Sheet 1: Estudios ----
    ws = wb.active
    ws.title = "Estudios"
    cols = ["Estudio", "DOI", "PMID", "Autores", "Año", "País", "Diseño", "Ámbito",
            "Financiación", "Registro", "Seguimiento", "Población", "N total", "Edad",
            "Sexo", "Inclusión", "Exclusión", "Brazos", "Datos solo en figuras", "Notas", "Fuente"]
    _header(ws, cols)
    for ext in extractions:
        d = ext.get("data") or {}
        st, pop = d.get("study", {}), d.get("population", {})
        arms = "; ".join(f"{a.get('name', '')} (n={a.get('n') if a.get('n') is not None else '?'})"
                         for a in d.get("arms", []))
        ws.append([d.get("label", ext.get("study_label", "")), ext.get("doi", ""), ext.get("pmid", ""),
                   st.get("authors", ""), st.get("year", ""), st.get("country", ""), st.get("design", ""),
                   st.get("setting", ""), st.get("funding", ""), st.get("registration", ""),
                   d.get("followup", ""), pop.get("description", ""), pop.get("total_n"),
                   pop.get("age", ""), pop.get("sex", ""), pop.get("key_inclusion", ""),
                   pop.get("key_exclusion", ""), arms, d.get("data_in_figures_only", ""),
                   d.get("notes", ""), ext.get("source", "")])
    _autosize(ws, {1: 24, 2: 20, 4: 22, 7: 18, 12: 30, 16: 28, 17: 28, 18: 26, 19: 30, 20: 26})

    # ---- Sheet 2: Desenlaces ----
    ws2 = wb.create_sheet("Desenlaces")
    cols2 = ["Estudio", "Desenlace", "Tipo", "Momento",
             "Int: eventos", "Int: N", "Int: media", "Int: DE", "Int: persona-tiempo",
             "Ctrl: eventos", "Ctrl: N", "Ctrl: media", "Ctrl: DE", "Ctrl: persona-tiempo",
             "Medida", "Efecto", "IC inf", "IC sup", "p", "Fuente", "Notas"]
    _header(ws2, cols2)
    for ext in extractions:
        d = ext.get("data") or {}
        label = d.get("label", ext.get("study_label", ""))
        for o in d.get("outcomes", []):
            iv, ct, eff = o.get("intervention", {}), o.get("control", {}), o.get("effect", {})
            ws2.append([label, o.get("name", ""), o.get("type", ""), o.get("timepoint", ""),
                        iv.get("events"), iv.get("n"), iv.get("mean"), iv.get("sd"), iv.get("person_time"),
                        ct.get("events"), ct.get("n"), ct.get("mean"), ct.get("sd"), ct.get("person_time"),
                        eff.get("measure", ""), eff.get("value"), eff.get("ci_lower"), eff.get("ci_upper"),
                        eff.get("p"), o.get("source", ""), o.get("notes", "")])
    _autosize(ws2, {1: 22, 2: 30, 3: 14, 15: 10, 20: 16, 21: 26})

    # ---- Sheet 3: Para meta-análisis ----
    ws3 = wb.create_sheet("Para meta-análisis")
    _header(ws3, _META_COLS)
    for ext in extractions:
        d = ext.get("data") or {}
        label = d.get("label", ext.get("study_label", ""))
        year = d.get("study", {}).get("year", "")
        for o in d.get("outcomes", []):
            row = _meta_row(label, o)
            row["year"] = year
            ws3.append([row.get(c, "") for c in _META_COLS])
    _autosize(ws3, {1: 22, 3: 28, 4: 14})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
