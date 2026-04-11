from __future__ import annotations

from pathlib import Path
import re


def _normalize(value: str | None) -> str:
    if value is None:
        return ''
    lowered = str(value).strip().lower()
    lowered = lowered.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
    return re.sub(r'[^a-z0-9]+', ' ', lowered).strip()


def fill_extraction_workbook(template_path: str | Path, studies: list[dict], output_path: str | Path) -> dict:
    from openpyxl import load_workbook

    template = Path(template_path)
    output = Path(output_path)
    wb = load_workbook(template)

    extraction_sheet = wb['Extraction Form'] if 'Extraction Form' in wb.sheetnames else wb.worksheets[0]
    characteristics = wb['Study Characteristics'] if 'Study Characteristics' in wb.sheetnames else None
    rob_sheet = wb['Risk of Bias'] if 'Risk of Bias' in wb.sheetnames else None

    variable_map = {
        'id del estudio': 'study_id',
        'primer autor': 'first_author',
        'ano de publicacion': 'year',
        'pais': 'country',
        'revista': 'journal',
        'diseno del estudio': 'design',
        'numero total de pacientes': 'n_total',
        'edad media de o mediana riq': 'age_mean',
        'edad media de o mediana r iq': 'age_mean',
        'consolidacion osea': 'consolidation',
        'tiempo hasta consolidacion': 'follow_up_months',
        'infeccion profunda postoperatoria': 'postop_infection',
        'seguimiento medio de o mediana riq': 'follow_up_months',
        'otras complicaciones': 'notes',
    }

    filled_sheets: list[str] = []
    for idx, study in enumerate(studies, start=1):
        if idx == 1:
            ws = extraction_sheet
            ws.title = f"Extraction Form - {study.get('study_id') or idx}"
        else:
            ws = wb.copy_worksheet(extraction_sheet)
            ws.title = f"Extraction Form - {study.get('study_id') or idx}"
        filled_sheets.append(ws.title)
        if study.get('title'):
            ws['A2'] = study.get('title')

        for row in range(1, ws.max_row + 1):
            variable = _normalize(ws.cell(row, 1).value)
            form_values = study.get('form_values') or {}
            if variable in form_values:
                ws.cell(row, 3).value = form_values[variable]
                continue
            key = variable_map.get(variable)
            if key:
                ws.cell(row, 3).value = study.get(key, 'No reportado')

    if characteristics is not None:
        headers = [_normalize(characteristics.cell(1, col).value) for col in range(1, characteristics.max_column + 1)]
        mapping = {
            'id': 'study_id',
            'autor': 'first_author',
            'ano': 'year',
            'pais': 'country',
            'diseno': 'design',
            'n total': 'n_total',
            'edad media': 'age_mean',
            'seguimiento meses': 'follow_up_months',
        }
        row_index = 2
        for study in studies:
            while row_index <= characteristics.max_row and any(characteristics.cell(row_index, col).value for col in range(1, characteristics.max_column + 1)):
                row_index += 1
            for col_index, header in enumerate(headers, start=1):
                key = mapping.get(header)
                if key:
                    characteristics.cell(row_index, col_index).value = study.get(key)
            row_index += 1

    if rob_sheet is not None:
        row_index = 2
        for study in studies:
            while row_index <= rob_sheet.max_row and any(rob_sheet.cell(row_index, col).value for col in range(1, rob_sheet.max_column + 1)):
                row_index += 1
            rob_sheet.cell(row_index, 1).value = study.get('study_id')
            row_index += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        'template_path': str(template),
        'output_path': str(output),
        'studies_written': len(studies),
        'filled_sheets': filled_sheets,
    }
